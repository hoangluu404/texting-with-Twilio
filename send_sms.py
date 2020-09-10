from twilio.rest import Client  # twilio plugins
import openpyxl
from os import system
import time

# get credentials
with open('Creds.txt') as file:
    account_sid = file.readline()
    auth_token = file.readline()
    my_phone = file.readline()

def restart():
    print('Continue in\n3 ')
    time.sleep(1)
    print('2 ')
    time.sleep(1)
    print('1')
    time.sleep(1)
    _ = system('cls')

def get_file():
    location = input("Please enter an Excel file or Excel location: ")
    return location

def get_message():
    message = input('Enter a message: ')
    return message

# get contacts inside excel file and send message
def send_message(user_choice):
    client = Client(account_sid, auth_token) # established connection with client
        
    if(user_choice == 1):
        number = '+1' + input('Enter a Phone Number: ')
        text = input('Enter a Message: ')
        message = client.messages \
                          .create(
                                 body = text+'\nLy Agency',
                                 from_= my_phone,
                                 #media_url=['https://www.lyagency.com/wp-content/uploads/2020/07/Health-icon-2.png'],
                                 to = number
                             )
        print('sent to ' + number + ': [' + text + '] with id '+ message.sid)
        print(message.price)
        #print(message.media._uri)
    elif(user_choice == 2):
        file_location = get_file()
        
        fileNotOpened = True
        while(fileNotOpened):
            try:
                wb = openpyxl.load_workbook(file_location)
                sheet = wb['Sheet1']
                fileNotOpened = False
            except:
                print("File does not exist")
                file_location = get_file()
                fileNotOpened = True

        text = get_message()
        for rows in range(2,sheet.max_row+1,1):
            if(sheet.cell(row=rows,column=3).value != 'done'):
            
                name = sheet.cell(row=rows, column=1).value
                phone_number = '+1'+ str(int(sheet.cell(row=rows, column=2).value))

                # send message to contacts not yet texted inside excel       
                message = client.messages \
                          .create(
                                 body = 'Hi '+name+',\n'+text+'\nLy Agency',
                                 from_= my_phone,
                                 #media_url=['https://www.lyagency.com/wp-content/uploads/2020/07/Health-icon-2.png'],
                                 to = phone_number
                             )
                print('sent to ' + name + ' ' + phone_number + ': [' + text + '] with id '+ message.sid)
                print(message.price)
        '''        print(message.media._uri)
                sheet.cell(row=rows,column=3).value = 'done'
        wb.save(file_location)'''

def main():
    user_choice = 1
    while(user_choice != 0):
        valid_choice = False
        while(valid_choice == False):
            try:
                user_choice = int(input('0 to exit\n1 to send 1 text to a number\n2 to send multiple text from Excel File\nEnter your Choice: '))
                valid_choice = True
            except:
                _ = system('cls')
                print('Invalid response!')
                valid_choice = False
        if(user_choice == 0):
            print('Exit')
        else:
            _ = system('cls')
            send_message(user_choice)
        if(user_choice != 0):
            restart()

if __name__ == "__main__":
    main()
